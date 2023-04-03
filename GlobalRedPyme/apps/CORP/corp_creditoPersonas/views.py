import io
from django.core.files.uploadedfile import InMemoryUploadedFile

from apps.CENTRAL.central_catalogo.models import Catalogo
from .models import CreditoPersonas, CodigoCredito
from apps.PERSONAS.personas_personas.models import Personas
from apps.CORP.corp_empresas.models import Empresas
from apps.PERSONAS.personas_personas.serializers import PersonasSearchSerializer
from .serializers import (
    CreditoPersonasSerializer, CreditoPersonasPersonaSerializer, CodigoCreditoSerializer
)
# Enviar Correo
from ...config.util import sendEmail
# Importar boto3
import boto3
import tempfile
import environ
# Publicar en sns
from .producer_ifi import publish
# Consumir en sqs
from .consumer_ifi import get_queue_url
from rest_framework import status
from rest_framework.response import Response
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from django.utils import timezone
from django.conf import settings
# Swagger
from drf_yasg import openapi
from drf_yasg.utils import swagger_auto_schema
# Lectura de AWS s3
import boto3
import re
from apps.config import config
# excel
import openpyxl
# Generar codigos aleatorios
import string
import random
# Sumar minutos
from dateutil.relativedelta import relativedelta
# ObjectId
from bson import ObjectId
# logs
from apps.CENTRAL.central_logs.methods import createLog, datosTipoLog, datosProductosMDP

# declaracion variables log
datosAux = datosProductosMDP()
datosTipoLogAux = datosTipoLog()
# asignacion datos modulo
logModulo = datosAux['modulo']
logApi = datosAux['api']
# asignacion tipo de datos
logTransaccion = datosTipoLogAux['transaccion']
logExcepcion = datosTipoLogAux['excepcion']


# CRUD
# CREAR
# 'methods' can be used to apply the same modification to multiple methods
@swagger_auto_schema(methods=['post'], request_body=CreditoPersonasSerializer)
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def creditoPersonas_create(request):
    timezone_now = timezone.localtime(timezone.now())
    logModel = {
        'endPoint': logApi + 'create/',
        'modulo': logModulo,
        'tipo': logExcepcion,
        'accion': 'CREAR',
        'fechaInicio': str(timezone_now),
        'dataEnviada': '{}',
        'fechaFin': str(timezone_now),
        'dataRecibida': '{}'
    }
    if request.method == 'POST':
        try:
            logModel['dataEnviada'] = str(request.data)
            request.data['created_at'] = str(timezone_now)
            if 'updated_at' in request.data:
                request.data.pop('updated_at')

            if 'nombres' in request.data:
                if request.data['nombres'] != "":
                    request.data['nombresCompleto'] = f"""{request.data['nombres']} {request.data['apellidos']}"""

            tipoCredito = 'Pymes-Normales'
            if 'tipoCredito' in request.data:
                if 'Pymes-PreAprobado' in request.data['tipoCredito']:
                    tipoCredito = request.data.pop('tipoCredito')

            serializer = CreditoPersonasSerializer(data=request.data)
            if serializer.is_valid():
                serializer.save()
                createLog(logModel, serializer.data, logTransaccion)
                if serializer.data['estado'] == 'Nuevo' and serializer.data['tipoCredito'] == 'Pymes-Normales':
                    # Publicar en la cola
                    publish(serializer.data)
                    enviarCorreoSolicitud(request.data['email'])
                if serializer.data['estado'] == 'Nuevo' and tipoCredito == 'Pymes-PreAprobado':
                    credito = serializer.data
                    credito['tipoCredito'] = 'Pymes-PreAprobado'
                    # Publicar en la cola
                    publish(credito)
                    enviarCorreoSolicitud(request.data['email'])
                return Response(serializer.data, status=status.HTTP_201_CREATED)
            createLog(logModel, serializer.errors, logExcepcion)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            err = {"error": 'Un error ha ocurrido: {}'.format(e)}
            createLog(logModel, err, logExcepcion)
            return Response(err, status=status.HTTP_400_BAD_REQUEST)


# ENCONTRAR UNO
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def creditoPersonas_listOne(request, pk):
    timezone_now = timezone.localtime(timezone.now())
    logModel = {
        'endPoint': logApi + 'listOne/',
        'modulo': logModulo,
        'tipo': logExcepcion,
        'accion': 'LEER',
        'fechaInicio': str(timezone_now),
        'dataEnviada': '{}',
        'fechaFin': str(timezone_now),
        'dataRecibida': '{}'
    }
    try:
        try:
            query = CreditoPersonas.objects.filter(pk=ObjectId(pk), state=1).first()
        except CreditoPersonas.DoesNotExist:
            err = {"error": "No existe"}
            createLog(logModel, err, logExcepcion)
            return Response(err, status=status.HTTP_404_NOT_FOUND)
        # tomar el dato
        if request.method == 'GET':
            serializer = CreditoPersonasSerializer(query)
            createLog(logModel, serializer.data, logTransaccion)
            return Response(serializer.data, status=status.HTTP_200_OK)
    except Exception as e:
        err = {"error": 'Un error ha ocurrido: {}'.format(e)}
        createLog(logModel, err, logExcepcion)
        return Response(err, status=status.HTTP_400_BAD_REQUEST)


# ACTUALIZAR
# 'methods' can be used to apply the same modification to multiple methods
@swagger_auto_schema(methods=['post'], request_body=CreditoPersonasSerializer)
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def creditoPersonas_update(request, pk):
    request.POST._mutable = True
    timezone_now = timezone.localtime(timezone.now())
    logModel = {
        'endPoint': logApi + 'update/',
        'modulo': logModulo,
        'tipo': logExcepcion,
        'accion': 'ESCRIBIR',
        'fechaInicio': str(timezone_now),
        'dataEnviada': '{}',
        'fechaFin': str(timezone_now),
        'dataRecibida': '{}'
    }
    try:
        try:
            logModel['dataEnviada'] = str(request.data)
            query = CreditoPersonas.objects.filter(pk=ObjectId(pk), state=1).order_by('-created_at').first()
        except CreditoPersonas.DoesNotExist:
            errorNoExiste = {'error': 'No existe'}
            createLog(logModel, errorNoExiste, logExcepcion)
            return Response(status=status.HTTP_404_NOT_FOUND)
        if request.method == 'POST':
            now = timezone.localtime(timezone.now())
            request.data['updated_at'] = str(now)
            if 'created_at' in request.data:
                request.data.pop('created_at')

            if 'claveFirma' in request.data:
                if request.data['claveFirma'] != '':
                    usuario = query.empresaInfo
                    date = now.strftime("D:%Y%m%d%H%M%S+00'00'")
                    dct = {
                        "aligned": 0,
                        "sigflags": 3,
                        "sigflagsft": 132,
                        "sigpage": 0,
                        "sigbutton": True,
                        "sigfield": "Signature1",
                        "auto_sigfield": True,
                        "sigandcertify": True,
                        "signaturebox": (470, 840, 570, 640),
                        "signature": usuario['reprsentante'],
                        # "signature_img": "signature_test.png",
                        "contact": usuario['correo'],
                        "location": "Ubicación",
                        "signingdate": date,
                        "reason": "Firmar documentos habilitantes",
                        "password": request.data['claveFirma'],
                    }
            if 'solicitudCreditoFirmado' in request.data:
                if request.data['solicitudCreditoFirmado'] is not None:
                    archivo_pdf_para_enviar_al_cliente = firmar(request, dct, 'solicitudCreditoFirmado')
                    request.data['solicitudCreditoFirmado'] = InMemoryUploadedFile(archivo_pdf_para_enviar_al_cliente,
                                                                                   'media',
                                                                                   'solicitudCreditoFirmado.pdf',
                                                                                   'application/pdf',
                                                                                   archivo_pdf_para_enviar_al_cliente.getbuffer().nbytes,
                                                                                   None
                                                                                   )
            if 'pagareFirmado' in request.data:
                if request.data['pagareFirmado'] is not None:
                    archivo_pdf_para_enviar_al_cliente = firmar(request, dct, 'pagareFirmado')
                    request.data['pagareFirmado'] = InMemoryUploadedFile(archivo_pdf_para_enviar_al_cliente,
                                                                         'media',
                                                                         'pagareFirmado.pdf',
                                                                         'application/pdf',
                                                                         archivo_pdf_para_enviar_al_cliente.getbuffer().nbytes,
                                                                         None
                                                                         )

            if 'contratosCuentaFirmado' in request.data:
                if request.data['contratosCuentaFirmado'] is not None:
                    archivo_pdf_para_enviar_al_cliente = firmar(request, dct, 'contratosCuentaFirmado')
                    request.data['contratosCuentaFirmado'] = InMemoryUploadedFile(archivo_pdf_para_enviar_al_cliente,
                                                                                  'media',
                                                                                  'contratosCuentaFirmado.pdf',
                                                                                  'application/pdf',
                                                                                  archivo_pdf_para_enviar_al_cliente.getbuffer().nbytes,
                                                                                  None
                                                                                  )
            if 'tablaAmortizacionFirmado' in request.data:
                if request.data['tablaAmortizacionFirmado'] is not None:
                    archivo_pdf_para_enviar_al_cliente = firmar(request, dct, 'tablaAmortizacionFirmado')
                    request.data['tablaAmortizacionFirmado'] = InMemoryUploadedFile(archivo_pdf_para_enviar_al_cliente,
                                                                                    'media',
                                                                                    'tablaAmortizacionFirmado.pdf',
                                                                                    'application/pdf',
                                                                                    archivo_pdf_para_enviar_al_cliente.getbuffer().nbytes,
                                                                                    None
                                                                                    )

            serializer = CreditoPersonasSerializer(query, data=request.data, partial=True)
            if serializer.is_valid():
                serializer.save()
                createLog(logModel, serializer.data, logTransaccion)
                usuario = serializer.data['user']
                email = usuario['email'] if usuario else serializer.data['email']
                if email == '' or email is None:
                    email = serializer.data['empresaInfo']['correo']
                if serializer.data['estado'] == 'Negado':
                    # Publicar en la cola
                    publish(serializer.data)
                    enviarCorreoNegado(serializer.data['montoLiquidar'], email)
                if serializer.data['estado'] == 'Por Completar':
                    # Publicar en la cola
                    publish(serializer.data)
                    enviarCorreoPorcompletar(serializer.data['montoLiquidar'], email)
                if serializer.data['estado'] == 'Aprobado':
                    if serializer.data['montoLiquidar']:
                        enviarCorreoAprobado(serializer.data['montoLiquidar'], email)
                if 'tipoCredito' in request.data:
                    if request.data['tipoCredito'] == '':
                        credito = serializer.data
                        credito['tipoCredito'] = 'Pymes-PreAprobado'
                        # Publicar en la cola
                        publish(credito)
                return Response(serializer.data)
            createLog(logModel, serializer.errors, logExcepcion)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    except Exception as e:
        err = {"error": 'Un error ha ocurrido: {}'.format(e)}
        createLog(logModel, err, logExcepcion)
        return Response(err, status=status.HTTP_400_BAD_REQUEST)


# ELIMINAR
@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def creditoPersonas_delete(request, pk):
    nowDate = timezone.localtime(timezone.now())
    logModel = {
        'endPoint': logApi + 'delete/',
        'modulo': logModulo,
        'tipo': logExcepcion,
        'accion': 'BORRAR',
        'fechaInicio': str(nowDate),
        'dataEnviada': '{}',
        'fechaFin': str(nowDate),
        'dataRecibida': '{}'
    }
    try:
        try:
            query = CreditoPersonas.objects.filter(user_id=pk, state=1).first()
        except CreditoPersonas.DoesNotExist:
            err = {"error": "No existe"}
            createLog(logModel, err, logExcepcion)
            return Response(err, status=status.HTTP_404_NOT_FOUND)
            return Response(status=status.HTTP_404_NOT_FOUND)
        # tomar el dato
        if request.method == 'DELETE':
            serializer = CreditoPersonasSerializer(query, data={'state': '0', 'updated_at': str(nowDate)}, partial=True)
            if serializer.is_valid():
                serializer.save()
                createLog(logModel, serializer.data, logTransaccion)
                return Response(serializer.data, status=status.HTTP_200_OK)
            createLog(logModel, serializer.errors, logExcepcion)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    except Exception as e:
        err = {"error": 'Un error ha ocurrido: {}'.format(e)}
        createLog(logModel, err, logExcepcion)
        return Response(err, status=status.HTTP_400_BAD_REQUEST)


# ENCONTRAR CODIGO CREDITO PREAPROBADO
@api_view(['POST'])
def creditoPersonas_creditoPreaprobado_codigo(request):
    timezone_now = timezone.localtime(timezone.now())
    logModel = {
        'endPoint': logApi + 'creditoPreaprobado/codigo',
        'modulo': logModulo,
        'tipo': logExcepcion,
        'accion': 'LEER',
        'fechaInicio': str(timezone_now),
        'dataEnviada': '{}',
        'fechaFin': str(timezone_now),
        'dataRecibida': '{}'
    }
    if request.method == 'POST':
        try:
            logModel['dataEnviada'] = str(request.data)
            # Filtros
            filters = {"state": "1"}

            if "codigo" in request.data:
                if request.data["codigo"] != '':
                    filters['codigoPreaprobado'] = request.data["codigo"]

            if "cedula" in request.data:
                if request.data["cedula"] != '':
                    filters['numeroIdentificacion'] = request.data["cedula"]

            if "rucEmpresa" in request.data:
                if request.data["rucEmpresa"] != '':
                    filters['rucEmpresa'] = request.data["rucEmpresa"]

            # Serializar los datos
            try:
                query = CreditoPersonas.objects.get(**filters)
            except CreditoPersonas.DoesNotExist:
                err = {"error": "No existe"}
                createLog(logModel, err, logExcepcion)
                return Response(err, status=status.HTTP_404_NOT_FOUND)

            # response = {'monto': query.monto, 'nombreCompleto': query.nombres + ' ' + query.apellidos,
            #             'tipoPersona': query.tipoPersona, 'estadoCivil': query.estadoCivil}
            # query.state = 0
            # query.save()

            # envio de datos
            serializer = CreditoPersonasSerializer(query)
            createLog(logModel, serializer.data, logTransaccion)
            return Response(serializer.data, status=status.HTTP_200_OK)
        except Exception as e:
            err = {"error": 'Un error ha ocurrido: {}'.format(e)}
            createLog(logModel, err, logExcepcion)
            return Response(err, status=status.HTTP_400_BAD_REQUEST)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def creditoPersonas_list(request):
    timezone_now = timezone.localtime(timezone.now())
    logModel = {
        'endPoint': logApi + 'list/',
        'modulo': logModulo,
        'tipo': logExcepcion,
        'accion': 'LEER',
        'fechaInicio': str(timezone_now),
        'dataEnviada': '{}',
        'fechaFin': str(timezone_now),
        'dataRecibida': '{}'
    }
    if request.method == 'POST':
        try:
            logModel['dataEnviada'] = str(request.data)
            # paginacion
            page_size = int(request.data['page_size'])
            page = int(request.data['page'])
            offset = page_size * page
            limit = offset + page_size
            # Filtros
            filters = {"state": "1"}

            if "empresaComercial_id" in request.data:
                if request.data["empresaComercial_id"] != '':
                    filters['empresaComercial_id'] = ObjectId(request.data["empresaComercial_id"])

            if "empresaIfis_id" in request.data:
                if request.data["empresaIfis_id"] != '':
                    filters['empresaIfis_id'] = ObjectId(request.data["empresaIfis_id"])

            if "estado" in request.data:
                if request.data["estado"] != '':
                    filters['estado__icontains'] = request.data["estado"]

            if "tipoCredito" in request.data:
                if request.data["tipoCredito"] != '':
                    filters['tipoCredito'] = str(request.data["tipoCredito"])

            if "user_id" in request.data:
                if request.data["user_id"] != '':
                    filters['user_id'] = str(request.data["user_id"])

            if "canal" in request.data:
                if request.data["canal"] != '':
                    filters['canal'] = str(request.data["canal"])

            if "numeroIdentificacion" in request.data:
                if request.data["numeroIdentificacion"] != '':
                    filters['numeroIdentificacion'] = str(request.data["numeroIdentificacion"])

            if "cargarOrigen" in request.data:
                if request.data["cargarOrigen"] != '':
                    filters['cargarOrigen'] = str(request.data["cargarOrigen"])

            # Serializar los datos
            query = CreditoPersonas.objects.filter(**filters).order_by('-created_at')
            serializer = CreditoPersonasSerializer(query[offset:limit], many=True)
            new_serializer_data = {'cont': query.count(),
                                   'info': serializer.data}
            # envio de datos
            return Response(new_serializer_data, status=status.HTTP_200_OK)
        except Exception as e:
            err = {"error": 'Un error ha ocurrido: {}'.format(e)}
            createLog(logModel, err, logExcepcion)
            return Response(err, status=status.HTTP_400_BAD_REQUEST)


# METODO SUBIR ARCHIVOS EXCEL
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def uploadEXCEL_creditosPreaprobados(request):
    contValidos = 0
    contInvalidos = 0
    contTotal = 0
    errores = []
    try:
        if request.method == 'POST':
            first = True  # si tiene encabezado
            uploaded_file = request.FILES['documento']
            # you may put validations here to check extension or file size
            wb = openpyxl.load_workbook(uploaded_file)
            # getting a particular sheet by name out of many sheets
            worksheet = wb["Clientes"]
            lines = list()
        for row in worksheet.iter_rows():
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))
            lines.append(row_data)

        for dato in lines:
            contTotal += 1
            if first:
                first = False
                continue
            else:
                if len(dato) == 7:
                    resultadoInsertar = insertarDato_creditoPreaprobado(dato, request.data['empresa_financiera'])
                    if resultadoInsertar != 'Dato insertado correctamente':
                        contInvalidos += 1
                        errores.append({"error": "Error en la línea " + str(contTotal) + ": " + str(resultadoInsertar)})
                    else:
                        contValidos += 1
                else:
                    contInvalidos += 1
                    errores.append({"error": "Error en la línea " + str(
                        contTotal) + ": la fila tiene un tamaño incorrecto (" + str(len(dato)) + ")"})

        result = {"mensaje": "La Importación se Realizo Correctamente",
                  "correctos": contValidos,
                  "incorrectos": contInvalidos,
                  "errores": errores
                  }
        return Response(result, status=status.HTTP_201_CREATED)

    except Exception as e:
        err = {"error": 'Error verifique el archivo, un error ha ocurrido: {}'.format(e)}
        return Response(err, status=status.HTTP_400_BAD_REQUEST)


# INSERTAR DATOS EN LA BASE INDIVIDUAL
def insertarDato_creditoPreaprobado(dato, empresa_financiera):
    try:
        timezone_now = timezone.localtime(timezone.now())
        data = {}
        data['vigencia'] = dato[0].replace('"', "")[0:10] if dato[0] != "NULL" else None
        data['concepto'] = dato[1].replace('"', "") if dato[1] != "NULL" else None
        data['monto'] = dato[2].replace('"', "") if dato[2] != "NULL" else None
        data['plazo'] = dato[3].replace('"', "") if dato[3] != "NULL" else None
        data['interes'] = dato[4].replace('"', "") if dato[4] != "NULL" else None
        data['estado'] = 'PreAprobado'
        data['tipoCredito'] = 'PreAprobado'
        data['canal'] = 'PreAprobado'
        persona = Personas.objects.filter(identificacion=dato[5], state=1).first()
        data['user_id'] = persona.user_id
        data['numeroIdentificacion'] = dato[5]
        data['nombres'] = persona.nombres
        data['apellidos'] = persona.apellidos
        data['nombresCompleto'] = persona.nombres + ' ' + persona.apellidos
        data['empresaIfis_id'] = empresa_financiera
        data['empresasAplican'] = dato[6]
        data['created_at'] = str(timezone_now)
        # inserto el dato con los campos requeridos
        CreditoPersonas.objects.create(**data)
        return 'Dato insertado correctamente'
    except Exception as e:
        return str(e)


# METODO SUBIR ARCHIVOS EXCEL
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def uploadEXCEL_creditosPreaprobados_empleados(request):
    contValidos = 0
    contInvalidos = 0
    contTotal = 0
    errores = []
    try:
        if request.method == 'POST':
            first = True  # si tiene encabezado
            uploaded_file = request.FILES['documento']
            # you may put validations here to check extension or file size
            wb = openpyxl.load_workbook(uploaded_file)
            # getting a particular sheet by name out of many sheets
            worksheet = wb["Clientes"]
            lines = list()
        for row in worksheet.iter_rows():
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))
            lines.append(row_data)

        for dato in lines:
            contTotal += 1
            if first:
                first = False
                continue
            else:
                if len(dato) == 11:
                    resultadoInsertar = insertarDato_creditoPreaprobado_empleado(dato,
                                                                                 request.data['empresa_financiera'])
                    if resultadoInsertar != 'Dato insertado correctamente':
                        contInvalidos += 1
                        errores.append({"error": "Error en la línea " + str(contTotal) + ": " + str(resultadoInsertar)})
                    else:
                        contValidos += 1
                else:
                    contInvalidos += 1
                    errores.append({"error": "Error en la línea " + str(
                        contTotal) + ": la fila tiene un tamaño incorrecto (" + str(len(dato)) + ")"})

        result = {"mensaje": "La Importación se Realizo Correctamente",
                  "correctos": contValidos,
                  "incorrectos": contInvalidos,
                  "errores": errores
                  }
        return Response(result, status=status.HTTP_201_CREATED)

    except Exception as e:
        err = {"error": 'Error verifique el archivo, un error ha ocurrido: {}'.format(e)}
        return Response(err, status=status.HTTP_400_BAD_REQUEST)


# INSERTAR DATOS EN LA BASE INDIVIDUAL
def insertarDato_creditoPreaprobado_empleado(dato, empresa_financiera):
    try:
        timezone_now = timezone.localtime(timezone.now())
        data = {}
        data['vigencia'] = dato[0].replace('"', "")[0:10] if dato[0] != "NULL" else None
        data['concepto'] = dato[1].replace('"', "") if dato[1] != "NULL" else None
        data['monto'] = dato[2].replace('"', "") if dato[2] != "NULL" else None
        data['plazo'] = dato[3].replace('"', "") if dato[3] != "NULL" else None
        data['interes'] = dato[4].replace('"', "") if dato[4] != "NULL" else None
        data['estado'] = 'PreAprobado'
        data['tipoCredito'] = 'Empleado'
        data['canal'] = 'Empleado'
        persona = Personas.objects.filter(identificacion=dato[5], state=1).first()
        data['user_id'] = persona.user_id
        data['numeroIdentificacion'] = dato[5]
        data['nombres'] = persona.nombres
        data['apellidos'] = persona.apellidos
        data['nombresCompleto'] = persona.nombres + ' ' + persona.apellidos
        data['empresaIfis_id'] = empresa_financiera
        data['empresasAplican'] = dato[10]
        data['created_at'] = str(timezone_now)
        # inserto el dato con los campos requeridos
        CreditoPersonas.objects.create(**data)
        return 'Dato insertado correctamente'
    except Exception as e:
        return str(e)


# ENCONTRAR UNO
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def creditoPersonas_listOne_persona(request, pk):
    timezone_now = timezone.localtime(timezone.now())
    logModel = {
        'endPoint': logApi + 'listOne/',
        'modulo': logModulo,
        'tipo': logExcepcion,
        'accion': 'LEER',
        'fechaInicio': str(timezone_now),
        'dataEnviada': '{}',
        'fechaFin': str(timezone_now),
        'dataRecibida': '{}'
    }
    try:
        try:
            query = CreditoPersonas.objects.filter(pk=ObjectId(pk), state=1).first()
        except CreditoPersonas.DoesNotExist:
            err = {"error": "No existe"}
            createLog(logModel, err, logExcepcion)
            return Response(err, status=status.HTTP_404_NOT_FOUND)
        # tomar el dato
        if request.method == 'GET':
            serializer = CreditoPersonasPersonaSerializer(query)
            createLog(logModel, serializer.data, logTransaccion)
            return Response(serializer.data, status=status.HTTP_200_OK)
    except Exception as e:
        err = {"error": 'Un error ha ocurrido: {}'.format(e)}
        createLog(logModel, err, logExcepcion)
        return Response(err, status=status.HTTP_400_BAD_REQUEST)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def creditoPersonas_listOne_usuario(request, pk):
    timezone_now = timezone.localtime(timezone.now())
    logModel = {
        'endPoint': logApi + 'listOne/usuario/' + pk,
        'modulo': logModulo,
        'tipo': logExcepcion,
        'accion': 'LEER',
        'fechaInicio': str(timezone_now),
        'dataEnviada': '{}',
        'fechaFin': str(timezone_now),
        'dataRecibida': '{}'
    }
    try:
        try:
            filters = {'state': 1}
            filters['user_id'] = pk
            filters['estado__icontains'] = request.data['estado']
            query = CreditoPersonas.objects.filter(**filters).first()
        except CreditoPersonas.DoesNotExist:
            err = {"error": "No existe"}
            createLog(logModel, err, logExcepcion)
            return Response(err, status=status.HTTP_404_NOT_FOUND)
        # tomar el dato
        if request.method == 'POST':
            serializer = CreditoPersonasSerializer(query)
            createLog(logModel, serializer.data, logTransaccion)
            return Response(serializer.data, status=status.HTTP_200_OK)
    except Exception as e:
        err = {"error": 'Un error ha ocurrido: {}'.format(e)}
        createLog(logModel, err, logExcepcion)
        return Response(err, status=status.HTTP_400_BAD_REQUEST)


@api_view(['GET'])
# @permission_classes([IsAuthenticated])
def creditoPersonas_lecturaArchivos(request, pk):
    timezone_now = timezone.localtime(timezone.now())
    logModel = {
        'endPoint': logApi + 'lecturaArchivos/' + pk,
        'modulo': logModulo,
        'tipo': logExcepcion,
        'accion': 'LEER',
        'fechaInicio': str(timezone_now),
        'dataEnviada': '{}',
        'fechaFin': str(timezone_now),
        'dataRecibida': '{}'
    }
    try:
        try:
            query = CreditoPersonas.objects.filter(pk=ObjectId(pk), state=1).first()
        except CreditoPersonas.DoesNotExist:
            err = {"error": "No existe"}
            createLog(logModel, err, logExcepcion)
            return Response(err, status=status.HTTP_404_NOT_FOUND)
        # tomar el dato
        if request.method == 'GET':
            print(query.identificacion.name)
            dato1 = None if query.identificacion.name is None else obtenerDatosArchivos(str(query.identificacion.name))
            dato2 = None if query.ruc.name is None else obtenerDatosArchivos(str(query.ruc.name))
            # serializer = CreditoPersonasPersonaSerializer(query)
            # createLog(logModel, serializer.data, logTransaccion)
            return Response({'cedula': dato1, 'ruc': dato2}, status=status.HTTP_200_OK)
    except Exception as e:
        err = {"error": 'Un error ha ocurrido: {}'.format(e)}
        createLog(logModel, err, logExcepcion)
        return Response(err, status=status.HTTP_400_BAD_REQUEST)


def obtenerDatosArchivos(nombreArchivo):
    # Function invokes
    jobId = InvokeTextDetectJob('globalredpymes', nombreArchivo)
    print("Started job with id: {}".format(jobId))
    respuesta = {}
    if (CheckJobComplete(jobId)):
        response = JobResults(jobId)
        for resultPage in response:
            for item in resultPage["Blocks"]:
                if item['BlockType'] == 'LINE':
                    if re.match("\d{10}001", item['Text']):
                        respuesta['ruc'] = item['Text']

                    elif re.match("No. \d{9}-[0-9]", item['Text']):
                        respuesta['identificacion'] = item['Text'][4:]

                    elif re.match("^\d{4}\-(0?[1-9]|1[012])\-(0?[1-9]|[12][0-9]|3[01])$", item['Text']):
                        respuesta['fechaExpiracion'] = item['Text']

                    elif re.match("[aA-Zz]\d{4}[aA-Zz]\d{4}", item['Text']):
                        respuesta['codigoDactilar'] = item['Text']

    print("-------------------Imprimir-----------------")
    return respuesta


import time


## Textract APIs used - "start_document_text_detection", "get_document_text_detection"
def InvokeTextDetectJob(bucket, nombreArchivo):
    response = None
    textarctmodule = boto3.client(
        'textract',
        aws_access_key_id=config.AWS_ACCESS_KEY_ID_TEXTRACT,
        aws_secret_access_key=config.AWS_SECRET_ACCESS_KEY_TEXTRACT,
        region_name='us-east-1'
    )
    response = textarctmodule.start_document_text_detection(
        DocumentLocation={
            'S3Object': {
                'Bucket': bucket,
                # 'Name': nombreArchivo
                'Name': 'CORP/documentosCreditosPersonas/62d97613bceeaa781e803920_1658498310065_comprobante_1.pdf'
            }
        }
    )
    return response["JobId"]


def CheckJobComplete(jobId):
    time.sleep(5)
    client = boto3.client(
        'textract',
        aws_access_key_id=config.AWS_ACCESS_KEY_ID_TEXTRACT,
        aws_secret_access_key=config.AWS_SECRET_ACCESS_KEY_TEXTRACT,
        region_name='us-east-1'
    )
    response = client.get_document_text_detection(JobId=jobId)
    status = response["JobStatus"]
    print("Job status: {}".format(status))
    while (status == "IN_PROGRESS"):
        time.sleep(5)
        response = client.get_document_text_detection(JobId=jobId)
        status = response["JobStatus"]
        print("Job status: {}".format(status))
    return status


def JobResults(jobId):
    pages = []
    client = boto3.client(
        'textract',
        aws_access_key_id=config.AWS_ACCESS_KEY_ID_TEXTRACT,
        aws_secret_access_key=config.AWS_SECRET_ACCESS_KEY_TEXTRACT,
        region_name='us-east-1'
    )
    response = client.get_document_text_detection(JobId=jobId)

    pages.append(response)
    print("Resultset page recieved: {}".format(len(pages)))
    nextToken = None
    if ('NextToken' in response):
        nextToken = response['NextToken']
        while (nextToken):
            response = client.get_document_text_detection(JobId=jobId, NextToken=nextToken)
            pages.append(response)
            print("Resultset page recieved: {}".format(len(pages)))
            nextToken = None
            if ('NextToken' in response):
                nextToken = response['NextToken']
    return pages


@api_view(['GET'])
def prueba(request):
    timezone_now = timezone.localtime(timezone.now())
    logModel = {
        'endPoint': logApi + 'listOne/',
        'modulo': logModulo,
        'tipo': logExcepcion,
        'accion': 'LEER',
        'fechaInicio': str(timezone_now),
        'dataEnviada': '{}',
        'fechaFin': str(timezone_now),
        'dataRecibida': '{}'
    }
    try:
        get_queue_url()
        msg = {"msg": "Se actualizo la cola"}
        return Response(msg, status=status.HTTP_202_ACCEPTED)
    except Exception as e:
        err = {"error": 'Un error ha ocurrido: {}'.format(e)}
        createLog(logModel, err, logExcepcion)
        return Response(err, status=status.HTTP_400_BAD_REQUEST)


def enviarCodigoCorreoMicroCredito(codigo, email):
    subject, from_email, to = 'Generación de código para crédito aprobado', "08d77fe1da-d09822@inbox.mailtrap.io", \
                              email
    txt_content = codigo
    html_content = f"""
                <html>
                    <body>
                        <h1>Su código de seguridad para consulta de créditos se ha generado</h1>
                        <br>
                        <p>Hola!</p>
                        <br>
                        <p>Su código de seguridad para consulta de créditos es: {codigo}</p>
                        <br>
                        Atentamente,
                        <br>
                        CrediCompra-Big Puntos.
                        <br>
                    </body>
                </html>
                """
    sendEmail(subject, txt_content, from_email, to, html_content)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def creditoPersonas_codigo_creditoAprobado(request):
    nowDate = timezone.localtime(timezone.now())
    logModel = {
        'endPoint': logApi + 'generar/codigo/creditoAprobado/',
        'modulo': logModulo,
        'tipo': logExcepcion,
        'accion': 'BORRAR',
        'fechaInicio': str(nowDate),
        'dataEnviada': '{}',
        'fechaFin': str(nowDate),
        'dataRecibida': '{}'
    }
    try:
        try:
            query = CreditoPersonas.objects.filter(numeroIdentificacion=request.data['numeroIdentificacion'], state=1,
                                                   estado='Aprobado').order_by('-created_at').first()
        except CreditoPersonas.DoesNotExist:
            err = {"error": "No existe"}
            createLog(logModel, err, logExcepcion)
            return Response(err, status=status.HTTP_404_NOT_FOUND)
        # tomar el dato
        if request.method == 'POST':
            # Genera el codigo
            longitud_codigo = Catalogo.objects.filter(tipo='CONFIG_TWILIO', nombre='LONGITUD_CODIGO',
                                                      state=1).first().valor
            codigo = (''.join(random.choice(string.digits) for _ in range(int(longitud_codigo))))
            enviarCodigoCorreoMicroCredito(codigo, query.user['email'])
            serializer = CodigoCreditoSerializer(data={'credito_id': str(query._id), 'codigo': codigo,
                                                       'numeroIdentificacion': request.data['numeroIdentificacion']})
            if serializer.is_valid():
                serializer.save()
                createLog(logModel, serializer.data, logTransaccion)
                return Response(serializer.data, status=status.HTTP_200_OK)
            createLog(logModel, serializer.errors, logExcepcion)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    except Exception as e:
        err = {"error": 'Un error ha ocurrido: {}'.format(e)}
        createLog(logModel, err, logExcepcion)
        return Response(err, status=status.HTTP_400_BAD_REQUEST)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def creditoPersonas_validar_codigo_creditoAprobado(request):
    nowDate = timezone.localtime(timezone.now())
    logModel = {
        'endPoint': logApi + 'validar/codigo/creditoAprobado',
        'modulo': logModulo,
        'tipo': logExcepcion,
        'accion': 'BORRAR',
        'fechaInicio': str(nowDate),
        'dataEnviada': '{}',
        'fechaFin': str(nowDate),
        'dataRecibida': '{}'
    }
    try:
        try:
            query = CodigoCredito.objects.filter(numeroIdentificacion=request.data['numeroIdentificacion'], state=1,
                                                 codigo=request.data['codigo']).first()
        except CodigoCredito.DoesNotExist:
            err = {"error": "No existe"}
            createLog(logModel, err, logExcepcion)
            return Response(err, status=status.HTTP_404_NOT_FOUND)
        # tomar el dato
        if request.method == 'POST':
            tiempo = Catalogo.objects.filter(tipo='CONFIG_DURACION_CREDITO_APROBADO',
                                             nombre='DURACION_CODIGO_CREDITO_APROBADO', state=1).first().valor
            duracion = query.created_at + relativedelta(minutes=int(tiempo))
            if duracion > timezone.now():
                credito = CreditoPersonas.objects.get(_id=ObjectId(query.credito_id))
                print(credito)
                serializer = CreditoPersonasSerializer(credito)
                createLog(logModel, serializer.data, logTransaccion)
                return Response(serializer.data, status=status.HTTP_200_OK)
            else:
                serializer = {'mensaje': 'No existe'}
                createLog(logModel, serializer, logExcepcion)
                return Response(serializer, status=status.HTTP_400_BAD_REQUEST)
    except Exception as e:
        err = {"error": 'Un error ha ocurrido: {}'.format(e)}
        createLog(logModel, err, logExcepcion)
        return Response(err, status=status.HTTP_400_BAD_REQUEST)


def enviarCorreoSolicitud(email):
    subject, from_email, to = 'Solicitud de Crédito Recibida – Crédito Pagos', "08d77fe1da-d09822@inbox.mailtrap.io", \
                              email
    txt_content = f"""
                        Global RedPyme - Crédito Pagos ha recibido su solicitud, estaremos en contacto con usted a la brevedad posible.
                            Crédito Pagos es la mejor opción para el crecimiento de su negocio
                        Atentamente,
                        Global RedPyme – Crédito Pagos
    """
    html_content = f"""
                <html>
                    <body>
                        <p>Global RedPyme - Crédito Pagos ha recibido su solicitud, estaremos en contacto con usted a la brevedad posible.</p>
                        <br>
                        <br>
                        <p><b>Crédito Pagos es la mejor opción para el crecimiento de su negocio</b></p>
                        <br>
                        Atentamente,
                        <br>
                        Global RedPyme – Crédito Pagos
                        <br>
                    </body>
                </html>
                """
    sendEmail(subject, txt_content, from_email, to, html_content)


from cryptography.hazmat import backends
from cryptography.hazmat.primitives.serialization import pkcs12
from endesive.pdf import cms


def firmar(request, dct, nombreArchivo):
    certificado = request.data['certificado']
    # environ init
    env = environ.Env()
    environ.Env.read_env()  # LEE ARCHIVO .ENV
    client_s3 = boto3.client(
        's3',
        aws_access_key_id=env.str('AWS_ACCESS_KEY_ID'),
        aws_secret_access_key=env.str('AWS_SECRET_ACCESS_KEY')
    )
    with tempfile.TemporaryDirectory() as d:
        ruta = d + 'SOLICITUD_REMATRICULA_DE_.pdf'
        s3 = boto3.resource('s3')
        archivo = s3.meta.client.download_file('globalredpymes', str(request.data[nombreArchivo]), ruta)
    contrasenia = request.data['claveFirma']
    p12 = pkcs12.load_key_and_certificates(
        certificado.read(), contrasenia.encode("ascii"), backends.default_backend()
    )
    datau = open(ruta, "rb").read()
    datas = cms.sign(datau, dct, p12[0], p12[1], p12[2], "sha256")
    archivo_pdf_para_enviar_al_cliente = io.BytesIO()
    archivo_pdf_para_enviar_al_cliente.write(datau)
    archivo_pdf_para_enviar_al_cliente.write(datas)
    archivo_pdf_para_enviar_al_cliente.seek(0)
    return archivo_pdf_para_enviar_al_cliente


def enviarCorreoNegado(montoAprobado, email):
    subject, from_email, to = 'Su solicitud de crédito de consumo ha sido NEGADO', "08d77fe1da-d09822@inbox.mailtrap.io", \
                              email
    txt_content = montoAprobado
    html_content = f"""
                <html>
                    <body>
                        <h1>¡LO SENTIMOS!</h1>
                        <br>
                        <h2>Su solicitud de crédito para realizar compras en las mejores Casas Comerciales con un crédito otorgado por una Cooperativa de Ahorro y Crédito regulada ha sido NEGADA.</h2>
                        <p>
                        Contáctese con su asesor a través de nuestro link de WhatsApp: <a href='https://wa.link/e8b3sa'>LINK</a>
                        </p>
                        <br>
                        Atentamente,
                        <br>
                        CrediCompra – Big Puntos
                        <br>
                    </body>
                </html>
                """
    sendEmail(subject, txt_content, from_email, to, html_content)


def enviarCorreoPorcompletar(montoAprobado, email):
    subject, from_email, to = 'DEVUELTA PARA COMPLETAR INFORMACIÓN', "08d77fe1da-d09822@inbox.mailtrap.io", \
                              email
    txt_content = montoAprobado
    html_content = f"""
                <html>
                    <body>
                        <h1>¡LO SENTIMOS!</h1>
                        <br>
                        <p>
                        Su solicitud de crédito para realizar compras en las mejores Casas Comerciales con un crédito 
                        otorgado por una Cooperativa de Ahorro y Crédito regulada ha sido DEVUELTA PARA COMPLETAR INFORMACIÓN.
                        </p>
                        <br>
                        <p>Contáctese con su asesor a través de nuestro link de WhatsApp: <a href='https://wa.link/szsyad'>LINK</a></p>
                        <br>
                        Atentamente,
                        <br>
                        CrediCompra – Big Puntos
                        <br>
                    </body>
                </html>
                """
    sendEmail(subject, txt_content, from_email, to, html_content)


def enviarCorreoAprobado(montoAprobado, email):
    subject, from_email, to = 'Su solicitud de crédito de consumo ha sido APROBADA', "08d77fe1da-d09822@inbox.mailtrap.io", \
                              email
    txt_content = montoAprobado
    html_content = f"""
                <html>
                    <body>
                        <h1>CRÉDITO APROBADO</h1>
                        <br>
                        <h2>Felicidades!</h2>
                        <p>
                        Su crédito para realizar compras en las mejores Casas Comerciales del país ha sido aprobado 
                        por un monto de {montoAprobado} ha sido aprobado. Para acceder a su crédito, 
                        siga los siguientes pasos:
                        </p>
                        <br>
                        <ol>
                             <li>Ingrese a <a href='https://credicompra.com/'>www.credicompra.com</a> y revise el catálogo de nuestras Casas Comerciales afiliadas.</li>
                             <li>Acérquese a la Casa Comercial de su preferencia y solicite realizar la compra con su crédito Aprobado.</li>
                             <li>Confirme sus datos</li>
                             <li>Escoja sus productos y listo. Pague con su Crédito Aprobado</li>
                        </ol>
                        <br>
                        <p>Si requiere asistencia personalizada, contáctenos a través del siguiente <a href='https://wa.link/5aips'>LINK</a></p>
                        <br>
                        Atentamente,
                        <br>
                        CrediCompra – Big Puntos
                        <br>
                    </body>
                </html>
                """
    sendEmail(subject, txt_content, from_email, to, html_content)

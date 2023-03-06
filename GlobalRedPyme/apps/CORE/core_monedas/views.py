from .models import Monedas
from apps.PERSONAS.personas_personas.models import Personas
from apps.CORP.corp_empresas.models import Empresas
from .serializers import (
    MonedasSerializer, MonedasUsuarioSerializer, ListMonedasSerializer, ListMonedasRegaladasSerializer
)
from rest_framework import status
from rest_framework.response import Response
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from django.utils import timezone
# Utils
from apps.utils import utils
# Enviar Correo
from apps.config.util import sendEmail
# Swagger
from drf_yasg import openapi
from drf_yasg.utils import swagger_auto_schema
# ObjectId
from bson import ObjectId
# excel
import openpyxl
# logs
from apps.CENTRAL.central_logs.methods import createLog, datosTipoLog, datosProductosMDP

# declaracion variables log
datosAux = datosProductosMDP()
datosTipoLogAux = datosTipoLog()
# asignacion datos modulo
logModulo = datosAux['modulo']
logApi = datosAux['api']
# asignacion tipo de datos
logTransaccion = datosTipoLogAux['transaccion']
logExcepcion = datosTipoLogAux['excepcion']


# CRUD CORE
# LISTAR TODOS
# 'methods' can be used to apply the same modification to multiple methods
@swagger_auto_schema(methods=['post'],
                     request_body=openapi.Schema(
                         type=openapi.TYPE_OBJECT,
                         required=['page_size', 'page'],
                         properties={
                             'page_size': openapi.Schema(type=openapi.TYPE_NUMBER),
                             'page': openapi.Schema(type=openapi.TYPE_NUMBER),
                             'user_id': openapi.Schema(type=openapi.TYPE_STRING),
                         },
                     ),
                     operation_description='Uninstall a version of Site',
                     responses={200: ListMonedasSerializer(many=True)})
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def monedas_list(request):
    timezone_now = timezone.localtime(timezone.now())
    logModel = {
        'endPoint': logApi + 'list/',
        'modulo': logModulo,
        'tipo': logExcepcion,
        'accion': 'LEER',
        'fechaInicio': str(timezone_now),
        'dataEnviada': '{}',
        'fechaFin': str(timezone_now),
        'dataRecibida': '{}'
    }
    if request.method == 'POST':
        try:
            logModel['dataEnviada'] = str(request.data)
            # paginacion
            page_size = int(request.data['page_size'])
            page = int(request.data['page'])
            offset = page_size * page
            limit = offset + page_size
            # Filtros
            filters = {"state": "1"}

            if 'user_id' in request.data:
                if request.data['user_id'] != '':
                    filters['user_id'] = request.data['user_id']

            # Serializar los datos
            query = Monedas.objects.filter(**filters).order_by('-created_at')
            serializer = ListMonedasSerializer(query[offset:limit], many=True)
            new_serializer_data = {'cont': query.count(),
                                   'info': serializer.data}
            # envio de datos
            return Response(new_serializer_data, status=status.HTTP_200_OK)
        except Exception as e:
            err = {"error": 'Un error ha ocurrido: {}'.format(e)}
            createLog(logModel, err, logExcepcion)
            return Response(err, status=status.HTTP_400_BAD_REQUEST)


# CREAR
# 'methods' can be used to apply the same modification to multiple methods
@swagger_auto_schema(methods=['post'], request_body=MonedasSerializer)
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def monedas_create(request):
    timezone_now = timezone.localtime(timezone.now())
    logModel = {
        'endPoint': logApi + 'create/',
        'modulo': logModulo,
        'tipo': logExcepcion,
        'accion': 'CREAR',
        'fechaInicio': str(timezone_now),
        'dataEnviada': '{}',
        'fechaFin': str(timezone_now),
        'dataRecibida': '{}'
    }
    if request.method == 'POST':
        try:
            logModel['dataEnviada'] = str(request.data)
            request.data['created_at'] = str(timezone_now)
            if 'updated_at' in request.data:
                request.data.pop('updated_at')

            serializer = MonedasSerializer(data=request.data)
            if serializer.is_valid():
                serializer.save()
                createLog(logModel, serializer.data, logTransaccion)
                return Response(serializer.data, status=status.HTTP_201_CREATED)
            createLog(logModel, serializer.errors, logExcepcion)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            err = {"error": 'Un error ha ocurrido: {}'.format(e)}
            createLog(logModel, err, logExcepcion)
            return Response(err, status=status.HTTP_400_BAD_REQUEST)


# ENCONTRAR UNO
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def monedas_listOne(request, pk):
    timezone_now = timezone.localtime(timezone.now())
    logModel = {
        'endPoint': logApi + 'listOne/',
        'modulo': logModulo,
        'tipo': logExcepcion,
        'accion': 'LEER',
        'fechaInicio': str(timezone_now),
        'dataEnviada': '{}',
        'fechaFin': str(timezone_now),
        'dataRecibida': '{}'
    }
    try:
        try:
            # Creo un ObjectoId porque la primaryKey de mongo es ObjectId
            pk = ObjectId(pk)
            query = Monedas.objects.get(pk=pk, state=1)
        except Monedas.DoesNotExist:
            err = {"error": "No existe"}
            createLog(logModel, err, logExcepcion)
            return Response(err, status=status.HTTP_404_NOT_FOUND)
        # tomar el dato
        if request.method == 'GET':
            serializer = MonedasSerializer(query)
            createLog(logModel, serializer.data, logTransaccion)
            return Response(serializer.data, status=status.HTTP_200_OK)
    except Exception as e:
        err = {"error": 'Un error ha ocurrido: {}'.format(e)}
        createLog(logModel, err, logExcepcion)
        return Response(err, status=status.HTTP_400_BAD_REQUEST)


# ACTUALIZAR
# 'methods' can be used to apply the same modification to multiple methods
@swagger_auto_schema(methods=['post'], request_body=MonedasSerializer)
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def monedas_update(request, pk):
    timezone_now = timezone.localtime(timezone.now())
    logModel = {
        'endPoint': logApi + 'update/',
        'modulo': logModulo,
        'tipo': logExcepcion,
        'accion': 'ESCRIBIR',
        'fechaInicio': str(timezone_now),
        'dataEnviada': '{}',
        'fechaFin': str(timezone_now),
        'dataRecibida': '{}'
    }
    try:
        try:
            logModel['dataEnviada'] = str(request.data)
            # Creo un ObjectoId porque la primaryKey de mongo es ObjectId
            pk = ObjectId(pk)
            query = Monedas.objects.get(pk=pk, state=1)
        except Monedas.DoesNotExist:
            errorNoExiste = {'error': 'No existe'}
            createLog(logModel, errorNoExiste, logExcepcion)
            return Response(status=status.HTTP_404_NOT_FOUND)
        if request.method == 'POST':
            now = timezone.localtime(timezone.now())
            request.data['updated_at'] = str(now)
            if 'created_at' in request.data:
                request.data.pop('created_at')

            serializer = MonedasSerializer(query, data=request.data, partial=True)
            if serializer.is_valid():
                serializer.save()
                createLog(logModel, serializer.data, logTransaccion)
                return Response(serializer.data)
            createLog(logModel, serializer.errors, logExcepcion)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    except Exception as e:
        err = {"error": 'Un error ha ocurrido: {}'.format(e)}
        createLog(logModel, err, logExcepcion)
        return Response(err, status=status.HTTP_400_BAD_REQUEST)

    # #ELIMINAR


@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def monedas_delete(request, pk):
    nowDate = timezone.localtime(timezone.now())
    logModel = {
        'endPoint': logApi + 'delete/',
        'modulo': logModulo,
        'tipo': logExcepcion,
        'accion': 'BORRAR',
        'fechaInicio': str(nowDate),
        'dataEnviada': '{}',
        'fechaFin': str(nowDate),
        'dataRecibida': '{}'
    }
    try:
        try:
            # Creo un ObjectoId porque la primaryKey de mongo es ObjectId
            pk = ObjectId(pk)
            persona = Monedas.objects.get(pk=pk, state=1)
        except Monedas.DoesNotExist:
            err = {"error": "No existe"}
            createLog(logModel, err, logExcepcion)
            return Response(err, status=status.HTTP_404_NOT_FOUND)
            return Response(status=status.HTTP_404_NOT_FOUND)
        # tomar el dato
        if request.method == 'DELETE':
            serializer = MonedasSerializer(persona, data={'state': '0', 'updated_at': str(nowDate)}, partial=True)
            if serializer.is_valid():
                serializer.save()
                createLog(logModel, serializer.data, logTransaccion)
                return Response(serializer.data, status=status.HTTP_200_OK)
            createLog(logModel, serializer.errors, logExcepcion)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    except Exception as e:
        err = {"error": 'Un error ha ocurrido: {}'.format(e)}
        createLog(logModel, err, logExcepcion)
        return Response(err, status=status.HTTP_400_BAD_REQUEST)

    # ENCONTRAR MONEDAS USUARIO


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def monedas_usuario(request, pk):
    timezone_now = timezone.localtime(timezone.now())
    logModel = {
        'endPoint': logApi + 'listOne/',
        'modulo': logModulo,
        'tipo': logExcepcion,
        'accion': 'LEER',
        'fechaInicio': str(timezone_now),
        'dataEnviada': '{}',
        'fechaFin': str(timezone_now),
        'dataRecibida': '{}'
    }
    try:
        try:
            query = Monedas.objects.filter(user_id=pk, state=1).order_by('-created_at').first()
        except Monedas.DoesNotExist:
            err = {"error": "No existe"}
            createLog(logModel, err, logExcepcion)
            return Response(err, status=status.HTTP_404_NOT_FOUND)
        # tomar el dato
        if request.method == 'GET':
            serializer = MonedasUsuarioSerializer(query)
            createLog(logModel, serializer.data, logTransaccion)
            return Response(serializer.data, status=status.HTTP_200_OK)
    except Exception as e:
        err = {"error": 'Un error ha ocurrido: {}'.format(e)}
        createLog(logModel, err, logExcepcion)
        return Response(err, status=status.HTTP_400_BAD_REQUEST)


# LISTAR MONEDAS OTORGADAS POR COMPRA
# 'methods' can be used to apply the same modification to multiple methods
@swagger_auto_schema(methods=['post'],
                     request_body=openapi.Schema(
                         type=openapi.TYPE_OBJECT,
                         required=['page_size', 'page'],
                         properties={
                             'page_size': openapi.Schema(type=openapi.TYPE_NUMBER),
                             'page': openapi.Schema(type=openapi.TYPE_NUMBER),
                             'user_id': openapi.Schema(type=openapi.TYPE_STRING),
                         },
                     ),
                     operation_description='Uninstall a version of Site',
                     responses={200: ListMonedasSerializer(many=True)})
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def monedas_listOtorgadas(request):
    timezone_now = timezone.localtime(timezone.now())
    logModel = {
        'endPoint': logApi + 'list/otorgadas/',
        'modulo': logModulo,
        'tipo': logExcepcion,
        'accion': 'LEER',
        'fechaInicio': str(timezone_now),
        'dataEnviada': '{}',
        'fechaFin': str(timezone_now),
        'dataRecibida': '{}'
    }
    if request.method == 'POST':
        try:
            logModel['dataEnviada'] = str(request.data)
            # paginacion
            page_size = int(request.data['page_size'])
            page = int(request.data['page'])
            offset = page_size * page
            limit = offset + page_size
            # Filtros
            filters = {"state": "1"}
            filters['tipo__in'] = ['Credito', 'Otro']

            if 'user_id' in request.data:
                if request.data['user_id'] != '':
                    filters['user_id'] = request.data['user_id']

            # Serializar los datos
            query = Monedas.objects.filter(**filters).order_by('-created_at')
            serializer = ListMonedasSerializer(query[offset:limit], many=True)
            new_serializer_data = {'cont': query.count(),
                                   'info': serializer.data}
            # envio de datos
            return Response(new_serializer_data, status=status.HTTP_200_OK)
        except Exception as e:
            err = {"error": 'Un error ha ocurrido: {}'.format(e)}
            createLog(logModel, err, logExcepcion)
            return Response(err, status=status.HTTP_400_BAD_REQUEST)


# METODO SUBIR ARCHIVOS EXCEL
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def uploadEXCEL_monedasRegaladas(request):
    contValidos = 0
    contInvalidos = 0
    contTotal = 0
    errores = []
    try:
        if request.method == 'POST':
            first = True  # si tiene encabezado
            uploaded_file = request.FILES['documento']
            # you may put validations here to check extension or file size
            wb = openpyxl.load_workbook(uploaded_file)
            # getting a particular sheet by name out of many sheets
            worksheet = wb["Empleados"]
            lines = list()
        for row in worksheet.iter_rows():
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))
            lines.append(row_data)

        for dato in lines:
            contTotal += 1
            if first:
                first = False
                continue
            else:
                if len(dato) == 9:
                    resultadoInsertar = insertarDato_monedas(dato)
                    if resultadoInsertar != 'Dato insertado correctamente':
                        contInvalidos += 1
                        errores.append({"error": "Error en la línea " + str(contTotal) + ": " + str(resultadoInsertar)})
                    else:
                        contValidos += 1
                else:
                    contInvalidos += 1
                    errores.append({"error": "Error en la línea " + str(
                        contTotal) + ": la fila tiene un tamaño incorrecto (" + str(len(dato)) + ")"})

        result = {"mensaje": "La Importación se Realizo Correctamente",
                  "correctos": contValidos,
                  "incorrectos": contInvalidos,
                  "errores": errores
                  }
        return Response(result, status=status.HTTP_201_CREATED)

    except Exception as e:
        err = {"error": 'Error verifique el archivo, un error ha ocurrido: {}'.format(e)}
        return Response(err, status=status.HTTP_400_BAD_REQUEST)


# INSERTAR DATOS EN LA BASE INDIVIDUAL
def insertarDato_monedas(dato):
    try:
        timezone_now = timezone.localtime(timezone.now())
        print(utils.__validar_ced_ruc(dato[1], 0))
        if (utils.__validar_ced_ruc(dato[1], 0) is False):
            return 'Cedula incorrecta'
        data = {}
        persona = Personas.objects.filter(identificacion=dato[1], state=1).first()
        if persona is not None:
            data['user_id'] = persona.user_id
        empresa = Empresas.objects.filter(ruc=dato[6], state=1).first()
        data['identificacion'] = dato[1]
        data['nombres'] = dato[2]
        data['apellidos'] = dato[3]
        data['email'] = dato[5]
        data['empresa_id'] = empresa._id
        data['tipo'] = 'Otro'
        data['estado'] = 'pendiente'
        data['credito'] = dato[6].replace('"', "") if dato[6] != "NULL" else None
        if persona is not None:
            monedas = Monedas.objects.filter(user_id=persona.user_id, state=1).order_by(
                '-created_at').first()
            if monedas is not None:
                monedasUsuario = monedas.saldo
            else:
                monedasUsuario = 0
        else:
            monedasUsuario = 0
        data['saldo'] = monedasUsuario + float(dato[4])
        data['descripcion'] = dato[8].replace('"', "") if dato[8] != "NULL" else None
        # data['fechaVigencia'] = dato[8].replace('"', "")[0:10] if dato[8] != "NULL" else None
        data['created_at'] = str(timezone_now)
        # inserto el dato con los campos requeridos
        Monedas.objects.create(**data)
        subject, from_email, to = f'Por su buen desempeño como empleado de {empresa.nombreComercial}, le premiamos', "08d77fe1da-d09822@inbox.mailtrap.io", \
                                  dato[5]
        txt_content = f"""
                        FELICIDADES!
                        USTED ACABA DE RECIBIR
                        {dato[4]} para que realice compras en los establecimientos afiliados pagando menos dinero en efectivo.
                        
                        Regístrese a través del portal https://portal.bigpuntos.com y reciba sus Big Puntos de recompensa en su cuenta
                        
                        Atentamente,
                        CrediCompra - Big Puntos
                """
        html_content = f"""
                <html>
                    <body>
                        <h1>FELICIDADES!</br>
                        <br>
                        <h1>USTED ACABA DE RECIBIR</h1>
                        <br>
                        <p>
                        {dato[4]} para que realice compras en los establecimientos afiliados pagando menos dinero en efectivo.
                        </p>
                        <br>
                        <p>Regístrese a través del portal https://portal.bigpuntos.com y reciba sus Big Puntos de recompensa en su cuenta</p>
                        <br>
                        Atentamente,
                        <br>
                        <b>CrediCompra - Big Puntos</b>
                        <br>
                    </body>
                </html>
                """
        sendEmail(subject, txt_content, from_email, to, html_content)
        return 'Dato insertado correctamente'
    except Exception as e:
        return str(e)


# METODO SUBIR ARCHIVOS EXCEL
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def uploadEXCEL_monedasRegaladasClientes(request):
    contValidos = 0
    contInvalidos = 0
    contTotal = 0
    errores = []
    try:
        if request.method == 'POST':
            first = True  # si tiene encabezado
            uploaded_file = request.FILES['documento']
            # you may put validations here to check extension or file size
            wb = openpyxl.load_workbook(uploaded_file)
            # getting a particular sheet by name out of many sheets
            worksheet = wb["Clientes"]
            lines = list()
        for row in worksheet.iter_rows():
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))
            lines.append(row_data)

        for dato in lines:
            contTotal += 1
            if first:
                first = False
                continue
            else:
                if len(dato) == 9:
                    resultadoInsertar = insertarDato_monedasClientes(dato)
                    if resultadoInsertar != 'Dato insertado correctamente':
                        contInvalidos += 1
                        errores.append({"error": "Error en la línea " + str(contTotal) + ": " + str(resultadoInsertar)})
                    else:
                        contValidos += 1
                else:
                    contInvalidos += 1
                    errores.append({"error": "Error en la línea " + str(
                        contTotal) + ": la fila tiene un tamaño incorrecto (" + str(len(dato)) + ")"})

        result = {"mensaje": "La Importación se Realizo Correctamente",
                  "correctos": contValidos,
                  "incorrectos": contInvalidos,
                  "errores": errores
                  }
        return Response(result, status=status.HTTP_201_CREATED)

    except Exception as e:
        err = {"error": 'Error verifique el archivo, un error ha ocurrido: {}'.format(e)}
        return Response(err, status=status.HTTP_400_BAD_REQUEST)


# INSERTAR DATOS EN LA BASE INDIVIDUAL
def insertarDato_monedasClientes(dato):
    try:
        timezone_now = timezone.localtime(timezone.now())
        print(utils.__validar_ced_ruc(dato[1], 0))
        if (utils.__validar_ced_ruc(dato[1], 0) is False):
            return 'Cedula incorrecta'
        data = {}
        persona = Personas.objects.filter(identificacion=dato[1], state=1).first()
        if persona is not None:
            data['user_id'] = persona.user_id
        empresa = Empresas.objects.filter(ruc=dato[6], state=1).first()
        data['identificacion'] = dato[1]
        data['nombres'] = dato[2]
        data['apellidos'] = dato[3]
        data['email'] = dato[5]
        data['empresa_id'] = empresa._id
        data['tipo'] = 'Otro'
        data['estado'] = 'pendiente'
        data['credito'] = dato[6].replace('"', "") if dato[6] != "NULL" else None
        if persona is not None:
            monedas = Monedas.objects.filter(user_id=persona.user_id, state=1).order_by(
                '-created_at').first()
            if monedas is not None:
                monedasUsuario = monedas.saldo
            else:
                monedasUsuario = 0
        else:
            monedasUsuario = 0
        data['saldo'] = monedasUsuario + float(dato[4])
        data['descripcion'] = dato[8].replace('"', "") if dato[8] != "NULL" else None
        # data['fechaVigencia'] = dato[8].replace('"', "")[0:10] if dato[8] != "NULL" else None
        data['created_at'] = str(timezone_now)
        # inserto el dato con los campos requeridos
        Monedas.objects.create(**data)
        subject, from_email, to = f'Por ser cliente de {empresa.nombreComercial}, le premiamos', "08d77fe1da-d09822@inbox.mailtrap.io", \
                                  dato[5]
        txt_content = f"""
                        {empresa.nombreComercial} LE PREMIA
                        Por ser cliente de {empresa.nombreComercial} le regalamos {dato[4]} para que realice compras de 
                        productos de varias categorías en establecimientos afiliados PAGANDO MENOS DINERO EN EFECTIVO.

                        Para acceder a sus Big Puntos, ingrese a: https://portal.bigpuntos.com/#/grp/login 

                        Si aún no tiene cuenta en Big Puntos, regístrese a través de: https://portal.bigpuntos.com/#/grp/registro y acceda a fabulosos premios.

                        Atentamente,
                        CrediCompra - Big Puntos
                """
        html_content = f"""
                <html>
                    <body>
                        <h1><b>{empresa.nombreComercial} LE PREMIA</b></h1>
                        <br>
                        <p>
                        Por ser cliente de {empresa.nombreComercial} le regalamos {dato[4]} para que realice compras de 
                        productos de varias categorías en establecimientos afiliados <b>PAGANDO MENOS DINERO EN EFECTIVO.</b>
                        </p>
                        <br>
                        <p>Para acceder a sus Big Puntos, ingrese a: https://portal.bigpuntos.com/#/grp/login</p>
                        <br>
                        <p>Si aún no tiene cuenta en Big Puntos, regístrese a través de: https://portal.bigpuntos.com/#/grp/registro y acceda a fabulosos premios.</p>
                        <br>
                        Atentamente,
                        <br>
                        <b>CrediCompra - Big Puntos</b>
                        <br>
                    </body>
                </html>
                """
        sendEmail(subject, txt_content, from_email, to, html_content)
        return 'Dato insertado correctamente'
    except Exception as e:
        return str(e)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def list_monedas_regaladas_empresa(request):
    timezone_now = timezone.localtime(timezone.now())
    logModel = {
        'endPoint': logApi + 'list/otorgadas/',
        'modulo': logModulo,
        'tipo': logExcepcion,
        'accion': 'LEER',
        'fechaInicio': str(timezone_now),
        'dataEnviada': '{}',
        'fechaFin': str(timezone_now),
        'dataRecibida': '{}'
    }
    if request.method == 'POST':
        try:
            logModel['dataEnviada'] = str(request.data)
            # paginacion
            page_size = int(request.data['page_size'])
            page = int(request.data['page'])
            offset = page_size * page
            limit = offset + page_size
            # Filtros
            filters = {"state": "1"}
            # filters['fechaVigencia__isnull'] = False
            # filters['fechaVigencia__gte'] = str(timezone_now)[0:10]

            if 'user_id' in request.data:
                if request.data['user_id'] != '':
                    filters['user_id'] = request.data['user_id']

            # Serializar los datos
            query = Monedas.objects.filter(**filters).order_by('-created_at')
            print(query)
            serializer = ListMonedasRegaladasSerializer(query[offset:limit], many=True)
            new_serializer_data = {'cont': query.count(),
                                   'info': serializer.data}
            # envio de datos
            return Response(new_serializer_data, status=status.HTTP_200_OK)
        except Exception as e:
            err = {"error": 'Un error ha ocurrido: {}'.format(e)}
            createLog(logModel, err, logExcepcion)
            return Response(err, status=status.HTTP_400_BAD_REQUEST)
